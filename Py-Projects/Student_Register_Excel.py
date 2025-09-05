import tkinter as tk
from tkinter import ttk
from tkinter import messagebox,ttk
from tkinter import *

#from tkcalendar import DateEntry

import os
import openpyxl


root=Tk()
root.geometry("1300x650")
root.title('Student Database')

top_frame=Frame(root,relief=GROOVE,borderwidth=9,height=1)
top_frame.pack(side='top')
head=Label(top_frame,text="STUDENT REGESTRATION MANAGEMENT SYSTEM",bg='red',fg='blue',height=1,width=177)
head.pack(fill=X)


#***********************************************
#       Store Data
#***********************************************



def store():
    PIN=pin.get()
    NAME=name.get()
    DOB=dob.get()
    BRANCH=branch.get()
    GENDER=gender.get()
    J_D=joindate.get()
    MOBILE=mobile.get()
    #AGE=age.get()
    
    
    if PIN and NAME and DOB and BRANCH and GENDER and J_D and MOBILE :

        filepath=r"C:\Users\admin\Desktop\Python\data.xlsx"
        
        if not os.path.exists(filepath):
            #create workbook file data.xlsx
            workbook=openpyxl.Workbook()
            #create a sheet 
            sheet=workbook.active
            # my sheet headings
            headings=["PIN","NAME","DOB","BRANCH","GENDER","JOIN DATE","MOBILE"]
            #insert my headings
            sheet.append(headings)
            #save file 
            workbook.save(filepath)
        #Means Already have file in this path
        #load my workbook in python
        workbook=openpyxl.load_workbook(filepath)
        root.title(f'Student Database - {filepath}')
        #activate my work book
        sheet=workbook.active
        #append My data in the xlsx file
        sheet.append([PIN,NAME,DOB,BRANCH,GENDER,J_D,MOBILE])
        #save my work book
        workbook.save(filepath)

        
        messagebox.showinfo(title="OK",message="Succesfully data stored")
        pin.delete(0,'end')
        name.delete(0,'end')
        mobile.delete(0,'end')
        gender.delete(0,'end')
        branch.delete(0,'end')
        display()

    else:
        messagebox.showwarning(title="Error",message="Cannot Fill Some Fields")
        #print('Error')





#***********************************************
#       Display data
#***********************************************



def display():
    filepath=r"C:\Users\admin\Desktop\Python\data.xlsx"
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook.active
    list_values=list(sheet.values)
    cols=list_values[0]


    tree=ttk.Treeview(root,column=cols,show="headings")

    for col_name in cols:
        tree.heading(col_name,text=col_name)
    
    
    tree.place(x=10,y=290)
    for values in list_values[1:]:
        tree.insert('',tk.END,values=values)








def modify():
    print("modify")
    MODIFY=Toplevel(root)
    MODIFY.geometry("500x400")
    MODIFY.title("Modify Student Record")
    MODIFY.configure(bg="#C5FFF6")
    MODIFY.resizable(False,False)
    heading=Label(MODIFY,text="Modify Student Record",width=23,font="sarif 15 bold",bg='blue',fg='white')
    heading.pack(fill=X)


    MODIFY.mainloop()


def search():
    print("search")
    SEARCH=Toplevel(root)
    SEARCH.geometry("500x400")
    SEARCH.title("Find Student Record")
    SEARCH.configure(bg="#D93C55")
    SEARCH.resizable(False,False)
    heading=Label(SEARCH,text="Search Student Record",width=23,font="sarif 15 bold",bg='blue',fg='white')
    heading.pack(fill=X)
    


    SEARCH.mainloop()

def delete():
    print("delete")
    DELETE=Toplevel(root)
    DELETE.geometry("400x300")
    DELETE.title("Delete Student Record")
    DELETE.configure(bg="#E4D9FF")
    DELETE.resizable(False,False)
    heading=Label(DELETE,text="Delete Student Record",width=23,font="sarif 15 bold",bg='blue',fg='white')
    heading.pack(fill=X)
    


    DELETE.mainloop()








#
#Data Entry Frame1
#
frame1=Frame(root,width=900,height=240,bg="blue")
frame1.place(x=10,y=40)

#
#Data Display Frame2
#


frame2=Frame(root,width=1260,height=320,bg="#E0FFAF")
frame2.place(x=10,y=290)

#
#Operations Entry Frame 3
#
frame3=Frame(root,width=360,height=240,bg="lightgreen")
frame3.place(x=910,y=40)

#
#==============================================================
#
heading1=Label(frame1,text="Student Data Entry",bg="red",fg='blue',width=30,font="Arial 14 bold")
heading1.place(x=280,y=10)


#
#Label and entry x and y
#

frame1_l_x=70
frame1_e_x=160
frame1_indicator_x=340

pin_l=Label(frame1,text='PIN :',width=10)
pin_l.place(x=frame1_l_x,y=60)
pin=Entry(frame1,width=20,font="arial 12")
pin.place(x=frame1_e_x,y=60)


name_l=Label(frame1,text='NAME :',width=10)
name_l.place(x=frame1_l_x,y=100)
name=Entry(frame1,width=20,font="arial 12")
name.place(x=frame1_e_x,y=100)



dob_l=Label(frame1,text='DOB :',width=10)
dob_l.place(x=frame1_l_x,y=140)
dob=Entry(frame1,width=18,font="arial 12")
dob.place(x=frame1_e_x,y=140)
#dob_l=Label(frame1,text=' <-- DD/MM/YYYY ',width=16)
#dob_l.place(x=frame1_indicator_x,y=140)

branch_l=Label(frame1,text='BRANCH :',width=10)
branch_l.place(x=frame1_l_x,y=180)
bra=['CME','ECE','EEE','AI']
branch=ttk.Combobox(root,values=bra,width=27)
branch.place(x=170,y=220)
#branch_l=Label(frame1,text=' <-- CME,ECE,EEE,AI ',width=16)
#branch_l.place(x=frame1_indicator_x,y=180)

#
#-----------------------------
#

gender_l=Label(frame1,text='GENDER :',width=10)
gender_l.place(x=400,y=60)
gen=['Male','Female','Others']
gender=ttk.Combobox(root,values=gen,width=27)
gender.place(x=505,y=100)
#gender_l=Label(frame1,text=' <-- M,F,O ',width=16)
#gender_l.place(x=frame1_indicator_x,y=220)

joindate_l=Label(frame1,text='JOIN DATE :',width=10)
joindate_l.place(x=400,y=100)
joindate=Entry(frame1,width=18,font="arial 12")
joindate.place(x=495,y=100)
#joindate_l=Label(frame1,text=' <-- DD/MM/YYYY ',width=16)
#joindate_l.place(x=frame1_indicator_x,y=260)


mobile_l=Label(frame1,text='MOBILE :',width=10)
mobile_l.place(x=400,y=140)
mobile=Entry(frame1,width=20,font="arial 12")
mobile.place(x=495,y=140)


#age_l=Label(frame1,text='AGE :',width=10)
#age_l.place(x=400,y=180)
#age=Entry(frame1,width=20,font="arial 12")
#age.place(x=495,y=180)




submit=Button(frame1,text="SAVE",command=store,width=10,bg='red',height=2 ,font="lusida 17 bold")
submit.place(x=710,y=94)



#============================================================

#Frame 2

#============================================================



























#============================================================

#Frame 3 Operation Buttons

#============================================================



modify=Button(frame3,text="MODIFY",command=modify,bg='blue',fg='white',height=2,width=10,font='arial 15')
modify.place(x=50,y=30)


search=Button(frame3,text="SEARCH",command=search,bg='orange',fg='white',height=2,width=10,font='arial 15')
search.place(x=200,y=30)

delete=Button(frame3,text="DELETE",command=delete,bg='red',fg='white',height=2,width=10,font='arial 15')
delete.place(x=50,y=130)

#display=Button(frame3,text="DISPLAY",command=display,bg='#DAFF75',fg='BLUE',height=2,width=10,font='arial 15')
#display.place(x=200,y=130)




 





























#
#************************************************************
#




menu=Menu(root)
root.config(menu=menu)
m1=Menu(menu,tearoff=False)
menu.add_cascade(label="Modify",menu=m1)
m1.add_command(label="Pin")
m1.add_command(label="Name")
m1.add_command(label="Gender")
m1.add_command(label="Branch")
m1.add_command(label="Mobile")




m2=Menu(menu,tearoff=False)
menu.add_cascade(label="Help",menu=m2)
m2.add_command(label="About")
m2.add_command(label="Conditions")
m2.add_command(label="Permissions")












#
#*************************************************************
#




























 


root.mainloop()
